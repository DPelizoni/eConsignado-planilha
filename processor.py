import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from validators import DataValidator
import io
import datetime

class FileProcessor:
    def __init__(self):
        self.logs = []
        self.output_filename = ""

    def process(self, input_file):
        self.logs = []
        try:
            # Read the origin spreadsheet
            df_origin = pd.read_excel(input_file)
            
            processed_data = []
            metadata = None
            
            for index, row in df_origin.iterrows():
                # Extract CNPJ and ensure it's a string with leading zeros if it's numeric
                cnpj_raw = str(row['numeroInscricaoEstabelecimento'])
                if cnpj_raw.replace('.','').replace('-','').isdigit():
                    cnpj_raw = cnpj_raw.replace('.','').replace('-','').zfill(14)
                
                # We no longer filter out rows based on 0001.
                # All rows are processed.
                
                data = {
                    'CNPJ': DataValidator.format_cnpj(cnpj_raw),
                    'TRABALHADOR': row['nomeTrabalhador'],
                    'CONTRATO': row['contrato'],
                    'DESCONTO INICIAL': row['competenciaInicioDesconto'],
                    'DESCONTO FINAL': row['competenciaFimDesconto'],
                    'PARCELAS': row['totalParcelas'],
                    'VALOR': float(row['valorParcela'])
                }
                processed_data.append(data)
                
                # Metadata logic:
                # B1 (Competência) and B2 (Razão Social) always from the first record.
                if metadata is None:
                    metadata = {
                        'razao_social': row['nomeEmpregador'],
                        'cnpj': "Informar o CNPJ da Matriz", # Default if no 0001 found
                        'competencia': row['competencia']
                    }
                
                # B3 (CNPJ): If we find a 0001, we prioritize it for this cell.
                if DataValidator.is_valid_cnpj_headquarters(cnpj_raw) and metadata['cnpj'] == "Informar o CNPJ da Matriz":
                    metadata['cnpj'] = DataValidator.format_cnpj(cnpj_raw)

            if not processed_data:
                return None, "A planilha de origem está vazia."
            
            df_final = pd.DataFrame(processed_data)
            
            # Prepare the output file
            output = io.BytesIO()
            template_path = "Planilha finalizada.xlsx"
            
            try:
                wb = openpyxl.load_workbook(template_path)
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            
            ws = wb.active
            # Metadata Writing (B1, B2, B3)
            competencia_b1 = ""
            if metadata['competencia']:
                comp_parts = str(metadata['competencia']).split('/')
                if len(comp_parts) == 2:
                    competencia_b1 = f"01/{comp_parts[0]}/{comp_parts[1]}"
                    # For filename: use mm-aaaa
                    mes_ano_file = f"{comp_parts[0]}-{comp_parts[1]}"
                else:
                    competencia_b1 = metadata['competencia']
                    # Fallback for filename if not in mm/aaaa format
                    mes_ano_file = str(metadata['competencia']).replace('/', '-')

            ws['B1'] = competencia_b1
            ws['B2'] = metadata['razao_social']
            ws['B3'] = metadata['cnpj']

            # Generate filename using the same logic as B1 (without the 01/)
            self.output_filename = f"{metadata['razao_social']}_eConsignado - Planilha_{mes_ano_file}.xlsx"

            # Writing data starting at row 7

            from copy import copy

            # Determine the maximum column to replicate based on the template row 7
            max_template_col = ws.max_column

            for r_idx, row_data in enumerate(dataframe_to_rows(df_final, index=False, header=False), 7):
                # 1. Write the mapped data (Columns 1 to N)
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Copy style from row 7 (source of truth for formatting)
                    if r_idx > 7:
                        source_cell = ws.cell(row=7, column=c_idx)
                        if source_cell.has_style:
                            cell.font = copy(source_cell.font)
                            cell.border = copy(source_cell.border)
                            cell.fill = copy(source_cell.fill)
                            cell.number_format = copy(source_cell.number_format)
                            cell.protection = copy(source_cell.protection)
                            cell.alignment = copy(source_cell.alignment)

                # 2. Replicate styles and formulas for EXTRA columns (like VALOR DESCONTADO, OBSERVAÇÕES)
                for c_idx in range(len(row_data) + 1, max_template_col + 1):
                    source_cell = ws.cell(row=7, column=c_idx)
                    new_cell = ws.cell(row=r_idx, column=c_idx)
                    
                    # Replicate Formula if it exists
                    if source_cell.data_type == 'f':
                        formula = source_cell.value
                        new_formula = formula.replace('7', str(r_idx))
                        new_cell.value = new_formula
                    
                    # ALWAYS replicate style for these columns if they exist in template
                    if r_idx > 7 and source_cell.has_style:
                        new_cell.font = copy(source_cell.font)
                        new_cell.border = copy(source_cell.border)
                        new_cell.fill = copy(source_cell.fill)
                        new_cell.number_format = copy(source_cell.number_format)
                        new_cell.protection = copy(source_cell.protection)
                        new_cell.alignment = copy(source_cell.alignment)
            
            wb.save(output)
            output.seek(0)
            
            return output, "Processamento concluído com sucesso. Todos os dados foram incluídos."

        except Exception as e:
            return None, f"Erro no processamento: {str(e)}"
