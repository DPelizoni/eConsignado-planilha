import pytest
from validators import DataValidator
from processor import FileProcessor
import pandas as pd
import io
import os
import openpyxl

# --- Tests for DataValidator ---

def test_cnpj_validator_headquarters():
    assert DataValidator.is_valid_cnpj_headquarters("55523356000100") is True
    assert DataValidator.is_valid_cnpj_headquarters("55.523.356/0001-00") is True
    assert DataValidator.is_valid_cnpj_headquarters("55523356000210") is False
    assert DataValidator.is_valid_cnpj_headquarters("invalid") is False

def test_cnpj_formatter():
    assert DataValidator.format_cnpj("55523356000100") == "55.523.356/0001-00"
    assert DataValidator.format_cnpj("123") == "123" # Too short, returns as is

def test_competencia_formatter():
    assert DataValidator.format_competencia("05/2026") == "01/05/2026"
    assert DataValidator.format_competencia(None) == ""

# --- Tests for FileProcessor ---

@pytest.fixture
def mock_excel_data():
    data = {
        'nomeEmpregador': ['EMPRESA TESTE', 'EMPRESA TESTE'],
        'numeroInscricaoEstabelecimento': ['12345678000199', '12345678000200'],
        'nomeTrabalhador': ['JOAO', 'MARIA'],
        'contrato': ['101', '102'],
        'competenciaInicioDesconto': ['05/2026', '06/2026'],
        'competenciaFimDesconto': ['05/2027', '06/2027'],
        'totalParcelas': [12, 12],
        'valorParcela': [100.50, 200.75],
        'competencia': ['05/2026', '05/2026']
    }
    df = pd.DataFrame(data)
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    return buffer

def test_processor_metadata_and_layout(mock_excel_data):
    processor = FileProcessor()
    # Note: This will look for "Planilha finalizada.xlsx". 
    # For testing, it should handle its absence gracefully.
    result, message = processor.process(mock_excel_data)
    
    assert result is not None
    assert "Processamento concluído" in message
    
    # Verify generated workbook
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    
    # Check B1, B2, B3
    assert ws['B1'].value == "01/05/2026"
    assert ws['B2'].value == "EMPRESA TESTE"
    assert ws['B3'].value == "12.345.678/0001-99" # Priority to 0001 found in data
    
    # Check data start at row 7
    # Row 7 Col 1 should be CNPJ
    assert ws.cell(row=7, column=1).value == "12.345.678/0001-99"
    # Row 8 Col 1 should be CNPJ of second row (even if not 0001)
    assert ws.cell(row=8, column=1).value == "12.345.678/0002-00"

def test_processor_no_headquarters():
    # Data with ONLY branches (no 0001)
    data = {
        'nomeEmpregador': ['FILIAL S/A'],
        'numeroInscricaoEstabelecimento': ['12345678000200'],
        'nomeTrabalhador': ['PEDRO'],
        'contrato': ['999'],
        'competenciaInicioDesconto': ['01/2026'],
        'competenciaFimDesconto': ['01/2027'],
        'totalParcelas': [10],
        'valorParcela': [50.00],
        'competencia': ['01/2026']
    }
    df = pd.DataFrame(data)
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    
    processor = FileProcessor()
    result, message = processor.process(buffer)
    
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    
    # B3 should show the specific message
    assert ws['B3'].value == "Informar o CNPJ da Matriz"
    # Row 7 should still contain the branch data
    assert ws.cell(row=7, column=1).value == "12.345.678/0002-00"

if __name__ == "__main__":
    pytest.main([__file__])
